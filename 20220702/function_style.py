from pprint import pprint
from openpyxl import Workbook, load_workbook


# 파일 생성
def createExcelFile(rows, filename: str) -> None:
  wb = Workbook()
  ws = wb.active

  for row in rows:
    ws.append(row)
  wb.save(filename)
  pprint(rows)
  

# 파일 읽기
def readExcelFile(filename: str) -> None:
  obj = load_workbook(filename)
  sheets = obj.sheetnames

  for sheet in sheets:
    rows = list(obj[sheet].values)
    pprint(rows)


if __name__ == "__main__":
  filename = "CHOEWY-R0.xlsx"
  createExcelFile([
    ['No.', 'Building', 'FireArea', 'FireArea Name'], 
    ['1', 'Reactor Building', 'CNB-000', 'Reactor building common Area'], 
    ['2', 'Primary Building', 'PB-014', 'Cable Spreding Room'],
    ['3', 'Turbine Building', 'TB-100', 'N-1E Battery Room']
  ], filename)
  readExcelFile(filename)